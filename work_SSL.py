from models.babyunet import BabyUnet
from models.models import get_model
from models.unet import Unet
from util import select_email_provider, send_email_to, show, plot_images, plot_tensors, create_montage
from skimage.morphology import disk
from skimage.filters import gaussian, median
from skimage.util import random_noise
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.transform import resize
from mask import Masker
from models.dncnn import DnCNN, work_DnCNN
from torch.nn import MSELoss
from torch.optim import Adam
from tqdm import tqdm
from PIL import Image

import numpy as np
import matplotlib.pyplot as plt
import cv2
import torch
import os 
import openpyxl
import shutil
import time 
import smtplib


class SSupervised(object) : 
        
    def __init__(self, params) : 
        plt.rc('figure', figsize = (5,5))
        self.p = params
        self.select_mode = params.selector
        self.SMTP = ''

        if params.selector == 'train' : 
            self.__set_train__()
        elif params.selector == 'test' : 
            self.__set_test__()
      
        # =============================================
        #   이미지 읽기, 그레이스케일 변환, 사이즈 조절 
        # =============================================
        self.img_gray = cv2.imread(os.path.join(self.p.img_path, self.p.img_file_name)) 
        self.img_cv_gray = cv2.cvtColor(self.img_gray, cv2.COLOR_BGR2GRAY)
        self.img_resize = resize(self.img_cv_gray, (self.p.resize,self.p.resize))
        # self.img_resize = cv2.equalizeHist(self.img_resize)
        # self.img_norm = cv2.normalize(self.img_resize, None, 0, 255, cv2.NORM_MINMAX)
        # print(f'self.img_norm : {self.img_norm} \n\n')
        
        
        # =============================================
        #   색상 전환, AHGN
        # =============================================
        flag = torch.Tensor(np.array([255]*self.p.resize*self.p.resize))
        self.fFlag = torch.FloatTensor(flag)
        self.fFlag = self.fFlag.view(self.img_resize.shape)
        fImg_dark = self.fFlag - self.img_resize 

        ahgn = torch.Tensor(np.array([127.5]*(self.p.resize**2)))
        ahgn = torch.FloatTensor(ahgn)
        ahgn = ahgn.view(fImg_dark.shape)

        self.minus_gaussian = fImg_dark - ahgn
        self.minus_gaussian = np.array(self.minus_gaussian)
        self.minus_gaussian = np.where(self.minus_gaussian < 0, 0, self.minus_gaussian)
        self.minus_gaussian = torch.Tensor(self.minus_gaussian)
        print(self.minus_gaussian)


        # =============================================
        #   노이즈 이미지 생성
        # =============================================
        # self.noisy_image = random_noise(self.img_resize, mode = self.p.noise_mode, var=1)        
        # self.noisy_image = random_noise(self.img_resize, mode = self.p.noise_mode)                # ...   0808 normalize image 
        # self.noisy_image = random_noise(self.img_norm, mode = self.p.noise_mode, var=0)
        
        # self.noisy_image = random_noise(self.img_norm, params = 'gaussian', var=3)        
        # self.noisy = torch.Tensor(self.noisy_image[np.newaxis, np.newaxis])         # ...     0728 test - without adding noise
        self.noisy = torch.Tensor(self.img_resize[np.newaxis, np.newaxis])
        # self.noisy = torch.Tensor(self.minus_gaussian[np.newaxis, np.newaxis])
        # self.noisy = torch.Tensor(self.img_resize)
        print(f'self.noisy.shape : {self.noisy.shape} \n\n')
        print(f'self.noisy : {self.noisy}\n\n')
        
        r'''
        # =============================================
        #   Minus Gaussian 
        # =============================================
        t = np.array([0.05]*640000)
        test_filter = torch.FloatTensor(t)
        test_filter = test_filter.view(self.noisy.shape)
        self.noisy = self.noisy - test_filter
        
        np.where(self.noisy < 0, 0, self.noisy)     # ...       0810 after minus avg point (0.5), fix each pixel value
        '''

        # =============================================
        #   Masking 
        # =============================================
        self.masker = Masker(width = 4, mode=self.p.mask_mode)
        
        self.my_address, self.my_password, self.recv_address = select_email_provider()

        if self.my_address != '' : 
            self.SMTP = smtplib.SMTP_SSL("smtp.gmail.com", 465)
            login_status, _ = self.SMTP.login(self.my_address, self.my_password)
            if login_status == 235 : 
                print(f'( V ) Authentication Success : {self.my_address}\n')
        
        # --------- GPU ---------
        device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
        print(f'\n( DEVICE ) : {device}')
        self.model = self.model.to(device)
        self.noisy = self.noisy.to(device)

    def __del__(self) : 
        if self.SMTP != '' : 
            self.SMTP.quit()
            print('\n\n** SMTP quit ! \n\n')

    def __set_train__(self) : 

        self.dir_title_by_date = ''
        self.record_train_time = "%02d-%02d" % (time.localtime().tm_hour, time.localtime().tm_min)
        self.excel_file_path = ''
        self.ckpt_file_path = ''
        self.target_excel = ''
        self.plt_image_title = ''
        self.first_cell_info = 10
        self.sheet_info_list = ['10', '5', '1']
        self.cell_info_dict = dict()
        self.cell_update_dict = dict()
        self.csv_first_cell = 'C'
        self.cell_EPOCH = self.csv_first_cell
        self.cell_PSNR = chr(ord(self.csv_first_cell)+1)
        self.cell_LOSS = chr(ord(self.csv_first_cell)+2)
        self.cell_VALID_LOSS = chr(ord(self.csv_first_cell)+3)
        self.losses = []
        self.val_losses = []
        self.best_val_loss = 1
        self.save_img_param = 0                                 # .. for save best image parameter
        self.learning_rate = self.p.lr
        self.epoch = self.p.epoch


        for sheet in self.sheet_info_list : 
            self.cell_info_dict[sheet] = self.first_cell_info - 1
            self.cell_update_dict[sheet] = 0
        self.model = work_DnCNN(1, num_of_layers = self.p.cnn_layer)
        # self.model = BabyUnet()
        # self.model = get_model("unet", 1, 1)
        sum(p.numel() for p in self.model.parameters() if p.requires_grad)


    r'''
    def __set_test__(self) : 
        # self.model = DnCNN(1, num_of_layers = self.p.cnn_layer)
        self.model = Unet(1, num_of_layers = self.p.cnn_layer)
        sum(p.numel() for p in self.model.parameters() if p.requires_grad)
        pass
    '''
    
    def work__train__(self) : 
        global plt
        
        self.best_val_loss = 1
        best_psnr, idx_loss, idx_val_loss = 0, 0, 0 
        self.ckpt_cnt_by_epoch = 0
        div_point = 100 
        self.psnr_ls = dict() 
        loss_function = MSELoss()
        optimizer = Adam(self.model.parameters(), 
                        lr=self.learning_rate
                        )
        
        i = 0
        last_output = ''
        for _ in range(self.epoch//div_point):
            target_dic = {'10': []}
            best_images = []
            ahgn_images = []

            r'''
            [ fix plt error ]
            >>> Fail to create pixmap with Tk_GetPixmap in TkImgPhotoInstanceSetSize

            [ Memo ]
            반복적으로 plt를 호출 하다보면 메모리를 초과하거나 과도한 반복으로 인해 에러가 발생한다. 
            plt 할당을 초기화 해주고, 메모리를 재지정 해준다. 
            '''
            if i % 2000 == 0 and i > 0: 
                del plt 
                from matplotlib import pyplot as plt 


            for _ in tqdm(range(div_point), desc="Train Process") :
                
                self.model.train()

                r'''
                loss_function(before, after)
                '''
                net_input, mask = self.masker.mask(self.noisy, i % (self.masker.n_masks - 1))
                net_output = self.model(net_input)
                
                r'''
                if last_output == '' : 
                    loss = loss_function(net_output*mask, self.noisy*mask)
                else : 
                    loss = loss_function(last_output, net_output*mask)
                '''

                loss = loss_function(net_output*mask, self.noisy*mask)
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
                
                if i % 10 == 0 : 
                    self.losses.append(loss.item())
                    self.model.eval()
                    
                    net_input, mask = self.masker.mask(self.noisy, self.masker.n_masks - 1)
                    net_output = self.model(net_input)

                    r'''
                    if last_output == '' : 
                        val_loss = loss_function(net_output*mask, self.noisy*mask)
                    else : 
                        val_loss = loss_function(last_output, net_output*mask)
                    '''

                    val_loss = loss_function(net_output*mask, self.noisy*mask)
                    self.val_losses.append(val_loss.item())
                    
                    idx_loss = round(loss.item(), 5)
                    idx_val_loss = round(val_loss.item(), 5)
                    denoised = np.clip(self.model(self.noisy).detach().cpu().numpy()[0, 0], 0, 1).astype(np.float64)

                    # check runtime error 
                    r'''
                    if i == 0 : 
                        denoised = np.clip(self.model(self.noisy).detach().cpu().numpy()[0, 0], 0, 1).astype(np.float64)
                    '''
                    if val_loss < self.best_val_loss:
                        self.best_val_loss = val_loss
                        # denoised = np.clip(self.model(self.noisy).detach().cpu().numpy()[0, 0], 0, 1).astype(np.float64)
                        # best_psnr = psnr(denoised, self.noisy_image)        # ...     0728 test - without adding noise
                        best_psnr = psnr(denoised, self.img_resize)                    
                        # self.best_images.append(denoised)                     # ...     0809 add all images, not only best cut 
                        if np.round(best_psnr, 5) not in self.psnr_ls.keys() : 
                            self.psnr_ls[np.round(best_psnr, 5)] = i
                    
                    best_images.append(denoised)
                    # best_images.append(self.fFlag - denoised)
                    # best_images.append(last_output)
                    # ahgn_images.append()


                    target_dic['10'].append([f"{i}/{self.epoch}", np.round(best_psnr, 5), idx_loss, idx_val_loss])
                    self.cell_update_dict['10'] += 1
                # ===========================================================
                r'''
                for div in target_dic.keys() : 
                    if i%int(div) == 0 : 
                        target_dic[div].append([f"{i}/{self.epoch}", np.round(best_psnr, 5), idx_loss, idx_val_loss])
                        self.cell_update_dict[div] += 1
                    else : 
                        break
                '''
                i += 1
                # self.work_save_model(self.model, i)       # ... save ckpt - test 
                time.sleep(0.05)

            # update information & check end of epoch 
            if i % div_point == 0 and i > 0 : 
                self.__save_csv__(target_dic)
                print(f"\n\n ● [ {i}/{self.epoch} ]", end = '')
                print(f"{'LOSS'.rjust(15, ' ')}{str(idx_loss).rjust(10, ' ')}{'VAL LOSS'.rjust(15, ' ')}{str(idx_val_loss).rjust(10, ' ')}")
                print('='.ljust(65, '='))
                self.__save__(i, best_images)
                self.__update_info__()
   

    # =============================================
    #   Save Image - Sequence
    # =============================================
    def __save__(self, _epoch, _images) : 
        assert 'images' in os.listdir(f'./results/{self.dir_title_by_date}/{self.record_train_time}'), f'\n\n** No such directory : " images "'

        # if len(self.best_images[self.save_img_param:len(self.best_images)]) > 1 : 
        if len(_images) > 1 : 
            # plot_images(self.best_images[self.save_img_param:len(self.best_images)])
            plot_images(_images)
            # self.save_img_param = len(self.best_images)

            savePath = f'./results/{self.dir_title_by_date}/{self.record_train_time}/images/sequence_{self.plt_image_title}-{_epoch}.png'
            plt.savefig(savePath)

            idx_subject = f"[ Training {self.plt_image_title}-{_epoch} ] → sequence_{self.plt_image_title}-{_epoch}.png"
            idx_data = f"[ EPOCH : {self.plt_image_title}-{_epoch} ]"
            if self.my_address != '' : 
                send_email_to(
                    _smtp = self.SMTP,
                    _my_address = self.my_address,
                    _subject = idx_subject,
                    _data = idx_data,
                    _img_path = savePath,
                    _recv_address = self.recv_address
                    )
            
            self.plt_image_title = _epoch
        
        # self.work_save_model(_epoch)       # ... save ckpt - test 


    # =============================================
    #   Make defualt directorys 
    # =============================================
    def __set_default_dirs__(self, _path = '.') : 
        now = time.localtime() 
        self.dir_title_by_date = "%04d-%02d-%02d" % (now.tm_year, now.tm_mon, now.tm_mday)

        if 'results' not in os.listdir(f'{_path}/') : 
            os.mkdir(f'{_path}/results')

        if self.dir_title_by_date not in os.listdir(f'{_path}/results') : 
            os.mkdir(f'{_path}/results/{self.dir_title_by_date}')

        assert 'result_format.xlsx' in os.listdir('./'), "\n\n** CSV result format not exists - check file name : result_format.xlsx"
        self.excel_file_path = f'./results/{self.dir_title_by_date}/{self.record_train_time}'

        if self.record_train_time not in os.listdir(f'./results/{self.dir_title_by_date}/') : 
            os.mkdir(f"{self.excel_file_path}")
            shutil.copy('./result_format.xlsx', f'{self.excel_file_path}/train_log.xlsx')
        
        if 'images' not in os.listdir(f'{self.excel_file_path}') : 
            os.mkdir(f'{self.excel_file_path}/images')

        if 'ckpt' not in os.listdir(f'./results/{self.dir_title_by_date}/{self.record_train_time}') :
            os.mkdir(f'./results/{self.dir_title_by_date}/{self.record_train_time}/ckpt')

    # =============================================
    #   CSV Update Information  
    # =============================================
    def __update_info__(self) : 
        if self.psnr_ls != {} : 
            print(f"{' ( Besat PSNR )'.ljust(20, ' ')}{'PSNR : '.rjust(20, ' ')}{str(max(self.psnr_ls.keys())).rjust(5, ' ')}{'EPOCH : '.rjust(15, ' ')}{str(self.psnr_ls.get(max(self.psnr_ls.keys()))).rjust(5, ' ')}")
        print(f" ○ {'Saved Check point : '.ljust(31, ' ')}{str(self.ckpt_cnt_by_epoch).rjust(30, ' ')}")        
        self.ckpt_cnt_by_epoch = 0 
        print(f" {'○ Sheet Title'.ljust(36,' ')}", end = '')
        for key, val in self.cell_update_dict.items() : 
            print(f"{key} ({val})".rjust(9,' '), end = '')
            self.cell_update_dict[key] = 0 
        print(f"\n ○ {'Save record at'.ljust(20, ' ')}{str(f'{self.target_excel}').rjust(40, ' ')}\n\n")
    

    # =============================================
    #   Save CSV
    # =============================================
    def __save_csv__(self, _save_target) : 
        if self.target_excel == '' : 
            self.target_excel = f'{self.excel_file_path}/train_log.xlsx'

        for _sheet_flag in _save_target.keys() : 
            wb = openpyxl.load_workbook(f'{self.target_excel}')
            for [_epoch, _psnr, _loss, _val_loss] in _save_target[_sheet_flag]: 

                self.cell_info_dict[_sheet_flag] += 1
                cell_point = self.cell_info_dict[_sheet_flag]
                
                ws = wb[_sheet_flag]
                ws[f'{self.cell_EPOCH}{cell_point}'] = _epoch
                ws[f'{self.cell_PSNR}{cell_point}'] = _psnr
                ws[f'{self.cell_LOSS}{cell_point}'] = _loss
                ws[f'{self.cell_VALID_LOSS}{cell_point}'] = _val_loss

            wb.save(f'{self.target_excel}')


    # =============================================
    #   Save Model - Test 
    # =============================================
    def work_save_model(self, _model, _epoch):
        CKPT_PATH = f'./results/{self.dir_title_by_date}/{self.record_train_time}/ckpt'

        fname = '{}/ssl-epoch{}.pt'.format(CKPT_PATH, _epoch)
        self.ckpt_cnt_by_epoch += 1
        # torch.save(self.model.state_dict(), fname)
        torch.save(_model.state_dict(), fname)


    # =============================================
    #   Load Model - Test 
    # =============================================
    def work_load_model(self, ckpt_fname):
        print('Loading checkpoint from: {}'.format(ckpt_fname))
        r'''
        if torch.cuda.is_available():
            self.model.load_state_dict(torch.load(ckpt_fname))
        else:
            self.model.load_state_dict(torch.load(ckpt_fname, map_location='cpu'))
        '''
        self.model.load_state_dict(torch.load(ckpt_fname, map_location='cpu'), strict=False)
        # self.model = torch.load(ckpt_fname)
        self.model.eval()

        self.work_test()

    def work_test(self):
        """Evaluates denoiser on test set."""

        # self.model.train(False)

        denoised = np.clip(self.model(self.noisy).detach().cpu().numpy()[0, 0], 0, 1).astype(np.float64)
        self.best_images.append(denoised)
        self.__save__(1)

        

    # =============================================
    #   Main
    # =============================================
    def __start__(self, save_file = None) :
        self.__set_default_dirs__()
        if self.select_mode == 'train' : 
            self.work__train__()
            # self.__save__()
        elif self.select_mode == 'test' : 
            self.work_test(save_file)
